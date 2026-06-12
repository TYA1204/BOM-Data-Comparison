import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from app.models import db

# ======================== 样式常量 ========================

# 通用
HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
NORMAL_FONT = Font(size=10)
BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT_WRAP = Alignment(horizontal='left', vertical='center', wrap_text=True)

# 层级填充色
ROOT_FILL = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')   # Root 淡橙
L1_FILL = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')   # L1 淡蓝
L2_FILL = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')   # L2 淡绿
L3_FILL = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')   # L3 白
L4_FILL = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')   # L4 浅灰
L5_FILL = PatternFill(start_color='FAFAFA', end_color='FAFAFA', fill_type='solid')   # L5 极浅灰

# 层级字体
L1_FONT = Font(bold=True, size=10, color='1F4E79')
L2_FONT = Font(bold=True, size=10, color='375623')
ROOT_FONT = Font(bold=True, size=10)

# 层级样式映射：level 1~5 → (fill, font)
LEVEL_STYLE = {
    1: (L1_FILL, L1_FONT),
    2: (L2_FILL, L2_FONT),
    3: (L3_FILL, NORMAL_FONT),
    4: (L4_FILL, NORMAL_FONT),
    5: (L5_FILL, NORMAL_FONT),
}

# 纵向单列列定义
BOM_COLUMNS = [
    ('序号', 6), ('层级', 6), ('物料编码', 26), ('物料名称', 35),
    ('用量', 8), ('单位', 6), ('位号', 18), ('ECN', 16),
]

# 差异报告样式
SEVERITY_STYLES = {
    'high': PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid'),
    'medium': PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'),
    'low': PatternFill(start_color='D9EAD3', end_color='D9EAD3', fill_type='solid'),
}
SEVERITY_LABELS = {'high': '高', 'medium': '中', 'low': '低'}
TYPE_LABELS = {
    'added': '新增物料', 'removed': '删除物料', 'modified': '变更物料',
    'version': '版本比对', 'cross_model': '跨机型比对',
}


# ======================== 辅助函数 ========================

def _fmt_qty(qty):
    """格式化数量：整数显示为int，小数保留原样。"""
    if qty is None or qty == '':
        return ''
    try:
        f = float(qty)
        if f == int(f):
            return str(int(f))
        return str(f)
    except (ValueError, TypeError):
        return str(qty)


def _clean_part_number(pn):
    """清理物料编码：去除前置的阿拉伯数字序号和顿号，如 '1、N030103...' → 'N030103...'"""
    import re
    return re.sub(r'^\s*\d+[、．]\s*', '', str(pn)).strip()


def _get_item_style(level):
    """根据层级返回 (fill, font)"""
    return LEVEL_STYLE.get(max(1, min(level, 5)), LEVEL_STYLE[3])


def _write_row(ws, row_idx, seq, level, pn, name, qty, unit, reference, version):
    """写入一行BOM数据（纵向单列格式）。"""
    fill, font = _get_item_style(level)
    level_label = 'Root' if level == 0 else f'L{level}'

    # A: 序号
    c = ws.cell(row=row_idx, column=1, value=seq)
    c.font = NORMAL_FONT; c.fill = fill; c.alignment = CENTER; c.border = BORDER

    # B: 层级
    c = ws.cell(row=row_idx, column=2, value=level_label)
    c.font = font; c.fill = fill; c.alignment = CENTER; c.border = BORDER

    # C: 物料编码（已清理）
    pn_clean = _clean_part_number(pn)
    c = ws.cell(row=row_idx, column=3, value=pn_clean)
    c.font = font; c.fill = fill; c.alignment = CENTER; c.border = BORDER

    # D: 物料名称
    c = ws.cell(row=row_idx, column=4, value=name)
    c.font = NORMAL_FONT; c.fill = fill; c.alignment = LEFT_WRAP; c.border = BORDER

    # E: 用量
    c = ws.cell(row=row_idx, column=5, value=_fmt_qty(qty))
    c.font = NORMAL_FONT; c.fill = fill; c.alignment = CENTER; c.border = BORDER

    # F: 单位
    c = ws.cell(row=row_idx, column=6, value=unit)
    c.font = NORMAL_FONT; c.fill = fill; c.alignment = CENTER; c.border = BORDER

    # G: 位号
    c = ws.cell(row=row_idx, column=7, value=reference)
    c.font = NORMAL_FONT; c.fill = fill; c.alignment = LEFT_WRAP; c.border = BORDER

    # H: ECN
    c = ws.cell(row=row_idx, column=8, value=version)
    c.font = NORMAL_FONT; c.fill = fill; c.alignment = CENTER; c.border = BORDER


def _export_items_vertical(ws, items, header_row):
    """将 items 按 line_no 顺序纵向展开写入工作表。

    直接按 line_no 顺序遍历，跳过 root 行（level=0），不递归。
    布局（纵向单列）：
      A: 序号 | B: 层级 | C: 物料编码（已清理）| D: 物料名称 |
      E: 用量 | F: 单位 | G: 位号 | H: ECN
    """
    row_idx = header_row
    seq = 0

    for it in items:
        level = it.get('level', 0)
        if level == 0:
            continue  # 跳过 root 行

        seq += 1
        pn = it['part_number']
        name = it.get('part_name', '')
        qty = it.get('quantity', '')
        unit = it.get('unit', '')
        reference = it.get('reference', '')
        version = it.get('version', '')

        _write_row(ws, row_idx, seq, level, pn, name, qty, unit, reference, version)
        row_idx += 1


# ======================== 公共导出接口 ========================

def generate_cleaned_bom_excel(bom_id):
    """导出单份清洗后的BOM数据为Excel（纵向单列结构化布局）。

    Args:
        bom_id: bom_header 的 id

    Returns:
        文件路径
    """
    header = db.query_one('SELECT * FROM bom_header WHERE id=?', (bom_id,))
    if not header:
        return None
    header = dict(header)

    items = db.query(
        'SELECT * FROM bom_item WHERE bom_id=? ORDER BY line_no',
        (bom_id,)
    )
    items = [dict(i) for i in items]
    if not items:
        return None

    wb = Workbook()
    sheet_name = header['bom_name'][:31] if header['bom_name'] else 'BOM数据'
    ws = wb.active
    ws.title = sheet_name

    # --- 信息头 ---
    info = [
        ('BOM名称', header['bom_name']),
        ('版本号', header.get('bom_version', '')),
        ('数据来源', header.get('source_file', '')),
        ('总行数', header.get('total_items', len(items))),
        ('导出时间', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
    ]
    for r, (k, v) in enumerate(info, 1):
        ws.cell(row=r, column=1, value=k).font = Font(bold=True, size=10)
        ws.cell(row=r, column=2, value=v).font = NORMAL_FONT

    header_row = len(info) + 2

    # --- 表头（纵向单列） ---
    for col_idx, (name, width) in enumerate(BOM_COLUMNS, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER
        ws.column_dimensions[cell.column_letter].width = width

    # --- 数据行（纵向单列） ---
    _export_items_vertical(ws, items, header_row + 1)

    # --- 保存 ---
    filename = f"BOM清洗数据_{header['bom_name']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    save_path = os.path.join(
        os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..')),
        'reports', filename
    )
    os.makedirs(os.path.dirname(save_path), exist_ok=True)
    wb.save(save_path)
    return save_path


def generate_cleaned_bom_dual_excel(bom_id_a, bom_id_b):
    """导出两份清洗后BOM数据到同一Excel（两个Sheet）。

    Returns:
        文件路径
    """
    header_a = db.query_one('SELECT * FROM bom_header WHERE id=?', (bom_id_a,))
    header_b = db.query_one('SELECT * FROM bom_header WHERE id=?', (bom_id_b,))
    if not header_a or not header_b:
        return None
    header_a, header_b = dict(header_a), dict(header_b)

    items_a = [dict(i) for i in db.query('SELECT * FROM bom_item WHERE bom_id=? ORDER BY line_no', (bom_id_a,))]
    items_b = [dict(i) for i in db.query('SELECT * FROM bom_item WHERE bom_id=? ORDER BY line_no', (bom_id_b,))]

    wb = Workbook()

    # --- 信息Sheet ---
    ws_info = wb.active
    ws_info.title = '概览'
    ws_info.column_dimensions['A'].width = 20
    ws_info.column_dimensions['B'].width = 40
    ws_info.column_dimensions['C'].width = 40

    info_header = ['项目', f"BOM-A：{header_a['bom_name']}", f"BOM-B：{header_b['bom_name']}"]
    for col_idx, h in enumerate(info_header, 1):
        cell = ws_info.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER

    info_rows = [
        ('数据来源', header_a.get('source_file', ''), header_b.get('source_file', '')),
        ('总行数', header_a.get('total_items', len(items_a)), header_b.get('total_items', len(items_b))),
        ('物料种类', len(set(i['part_number'] for i in items_a)), len(set(i['part_number'] for i in items_b))),
        ('导出时间', datetime.now().strftime('%Y-%m-%d %H:%M:%S'), ''),
    ]
    for r, (k, va, vb) in enumerate(info_rows, 2):
        ws_info.cell(row=r, column=1, value=k).font = Font(bold=True, size=10)
        ws_info.cell(row=r, column=2, value=va).font = NORMAL_FONT
        ws_info.cell(row=r, column=3, value=vb).font = NORMAL_FONT

    # --- 两个数据Sheet（纵向单列） ---
    for header, items, sheet_tag in [
        (header_a, items_a, 'A'),
        (header_b, items_b, 'B'),
    ]:
        ws = wb.create_sheet(title=f"BOM-{sheet_tag} {header['bom_name'][:20]}")
        # 信息头
        info = [
            ('BOM名称', header['bom_name']),
            ('版本号', header.get('bom_version', '')),
            ('数据来源', header.get('source_file', '')),
            ('总行数', header.get('total_items', len(items))),
            ('导出时间', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
        ]
        for r, (k, v) in enumerate(info, 1):
            ws.cell(row=r, column=1, value=k).font = Font(bold=True, size=10)
            ws.cell(row=r, column=2, value=v).font = NORMAL_FONT
        header_row = len(info) + 2
        # 表头
        for col_idx, (name, width) in enumerate(BOM_COLUMNS, 1):
            cell = ws.cell(row=header_row, column=col_idx, value=name)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER
            cell.border = BORDER
            ws.column_dimensions[cell.column_letter].width = width
        # 数据
        _export_items_vertical(ws, items, header_row + 1)

    # --- 保存 ---
    filename = f"BOM对比_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    save_path = os.path.join(
        os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..')),
        'reports', filename
    )
    os.makedirs(os.path.dirname(save_path), exist_ok=True)
    wb.save(save_path)
    return save_path


# ======================== 按组件导出 ========================

def generate_components_export_excel(bom_id, component_pns):
    """按选中的组件导出Excel（每个组件一个Sheet，包含其所有子件）。

    Args:
        bom_id: BOM ID
        component_pns: 组件物料号列表（单位=ST）

    Returns:
        文件路径
    """
    header = db.query_one('SELECT * FROM bom_header WHERE id=?', (bom_id,))
    if not header:
        return None
    header = dict(header)

    # 获取所有相关物料（组件本身 + 其所有子件）
    all_pns = set(component_pns)

    def collect_children(parent_pn):
        children = db.query(
            'SELECT part_number FROM bom_item WHERE bom_id=? AND parent_pn=?',
            (bom_id, parent_pn)
        )
        for c in children:
            pn = c['part_number']
            if pn not in all_pns:
                all_pns.add(pn)
                collect_children(pn)

    for pn in component_pns:
        collect_children(pn)

    # 批量查询所有相关物料
    placeholders = ','.join(['?'] * len(all_pns))
    items = db.query(
        f'SELECT * FROM bom_item WHERE bom_id=? AND part_number IN ({placeholders}) ORDER BY line_no',
        [bom_id] + list(all_pns)
    )
    items = [dict(i) for i in items]

    if not items:
        return None

    wb = Workbook()
    wb.remove(wb.active)

    # 按组件分组
    comp_map = {pn: [] for pn in component_pns}
    for it in items:
        ppn = it.get('parent_pn', '')
        if ppn in comp_map:
            comp_map[ppn].append(it)
        elif it['part_number'] in comp_map:
            pass  # 组件本身
        else:
            # 子件的子件，找到其顶层组件
            for cpn in component_pns:
                if _is_descendant_of(it['part_number'], cpn, bom_id):
                    comp_map[cpn].append(it)
                    break

    # 每个组件一个Sheet
    for idx, pn in enumerate(component_pns):
        comp_items = [i for i in items if i['part_number'] == pn]
        children = comp_map.get(pn, [])
        sheet_items = comp_items + sorted(children, key=lambda x: x['line_no'])

        comp_name = comp_items[0]['part_name'] if comp_items else pn

        # PC子件向上归并为父级ST组件名称
        # 构建 item 查找表：part_number → item dict
        item_lookup = {it['part_number']: it for it in sheet_items}

        def find_st_ancestor_name(pn_to_find):
            """向上追溯到最近的 ST 祖先组件，返回其名称。"""
            current = pn_to_find
            visited = set()
            while current in item_lookup:
                item = item_lookup[current]
                if item['unit'] == 'ST':
                    return item.get('part_name', '')
                parent = item.get('parent_pn', '')
                if not parent or parent in visited:
                    break
                visited.add(parent)
                current = parent
            return ''

        for it in sheet_items:
            if it['unit'] == 'ST':
                continue  # ST组件本身保持不变
            st_name = find_st_ancestor_name(it.get('parent_pn', ''))
            if st_name:
                it['part_name'] = st_name

        sheet_title = f"{pn[:15]}" if len(pn) > 15 else pn
        ws = wb.create_sheet(title=sheet_title)

        # 信息头
        ws.cell(row=1, column=1, value='组件').font = Font(bold=True, size=10)
        ws.cell(row=1, column=2, value=f"{pn} {comp_name}").font = NORMAL_FONT
        ws.cell(row=2, column=1, value='子件数量').font = Font(bold=True, size=10)
        ws.cell(row=2, column=2, value=len(children)).font = NORMAL_FONT

        header_row = 4
        for col_idx, (name, width) in enumerate(BOM_COLUMNS, 1):
            cell = ws.cell(row=header_row, column=col_idx, value=name)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER
            cell.border = BORDER
            ws.column_dimensions[cell.column_letter].width = width

        _export_items_vertical(ws, sheet_items, header_row + 1)

    # 保存
    filename = f"BOM组件导出_{header['bom_name']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    save_path = os.path.join(
        os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..')),
        'reports', filename
    )
    os.makedirs(os.path.dirname(save_path), exist_ok=True)
    wb.save(save_path)
    return save_path


def _is_descendant_of(part_number, ancestor_pn, bom_id, visited=None):
    """判断 part_number 是否是 ancestor_pn 的后代（递归检查parent链）。"""
    if visited is None:
        visited = set()
    if part_number in visited:
        return False
    visited.add(part_number)
    row = db.query_one(
        'SELECT parent_pn FROM bom_item WHERE bom_id=? AND part_number=? LIMIT 1',
        (bom_id, part_number)
    )
    if not row:
        return False
    parent = row['parent_pn']
    if parent == ancestor_pn:
        return True
    return _is_descendant_of(parent, ancestor_pn, bom_id, visited)


# ======================== 差异比对报告 ========================

def generate_excel_report(task_id):
    """Generate Excel comparison report.

    Returns file path.
    """
    task = db.query_one('SELECT * FROM comparison_task WHERE id=?', (task_id,))
    if not task:
        return None

    results = db.query(
        'SELECT * FROM comparison_result WHERE task_id=? ORDER BY severity DESC, diff_type, id',
        (task_id,)
    )

    wb = Workbook()

    # --- 汇总Sheet ---
    ws_summary = wb.active
    ws_summary.title = '汇总'

    stats = db.query('''
        SELECT diff_type, severity, COUNT(*) as cnt
        FROM comparison_result WHERE task_id=?
        GROUP BY diff_type, severity ORDER BY diff_type
    ''', (task_id,))

    td = dict(task)
    # 获取基准/目标BOM名称
    source_bom = db.query_one('SELECT bom_name, bom_version FROM bom_header WHERE id=?', (td.get('source_bom_id', 0),))
    target_bom = db.query_one('SELECT bom_name, bom_version FROM bom_header WHERE id=?', (td.get('target_bom_id', 0),))

    summary_data = [
        ('BOM 差异比对报告', ''),
        ('', ''),
        ('任务名称', td.get('task_name', '')),
        ('基准 BOM', f"{source_bom['bom_name'] if source_bom else '?'}  ({source_bom['bom_version'] if source_bom else '?'})"),
        ('目标 BOM', f"{target_bom['bom_name'] if target_bom else '?'}  ({target_bom['bom_version'] if target_bom else '?'})"),
        ('比对类型', TYPE_LABELS.get(td.get('comparison_type', ''), td.get('comparison_type', ''))),
        ('创建时间', td.get('created_at', '')),
        ('完成时间', td.get('completed_at', '')),
        ('', ''),
        ('--- 统计数据 ---', ''),
    ]

    total = 0
    for s in stats:
        type_label = TYPE_LABELS.get(s['diff_type'], s['diff_type'])
        sev_label = SEVERITY_LABELS.get(s['severity'], s['severity'])
        summary_data.append((f'  {type_label}（严重度：{sev_label}）', s['cnt']))
        total += s['cnt']
    summary_data.append(('', ''))
    summary_data.append(('差异总数', total))

    for row_idx, (key, val) in enumerate(summary_data, 1):
        ws_summary.cell(row=row_idx, column=1, value=key).font = NORMAL_FONT
        ws_summary.cell(row=row_idx, column=2, value=val).font = NORMAL_FONT

    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 40

    # --- 明细Sheet ---
    ws_detail = wb.create_sheet('明细')

    # 表头：基准BOM=来源(原值A)，目标BOM=对比(新值B)
    headers = ['序号', '差异类型', '分类', '严重度',
               '物料号(基准BOM)', '物料号(目标BOM)',
               '物料名称(基准BOM)', '物料名称(目标BOM)',
               '变更字段', '基准BOM值', '目标BOM值',
               '位号(基准BOM)', '位号(目标BOM)',
               '用量(基准BOM)', '用量(目标BOM)', '差异量(+/-)', '匹配度']

    for col_idx, h in enumerate(headers, 1):
        cell = ws_detail.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER

    for row_idx, r in enumerate(results, 2):
        diff_type_cn = TYPE_LABELS.get(r['diff_type'], r['diff_type'])
        severity_cn = SEVERITY_LABELS.get(r['severity'], r['severity'])
        qty_a = float(r['quantity_a'] or 0)
        qty_b = float(r['quantity_b'] or 0)
        diff_qty = qty_b - qty_a  # 目标BOM - 基准BOM，正值=增加，负值=减少
        row_data = [
            row_idx - 1, diff_type_cn, r['diff_category'], severity_cn,
            r['part_number_a'], r['part_number_b'],
            r['part_name_a'], r['part_name_b'],
            r['field_name'], r['old_value'], r['new_value'],
            r['reference_a'], r['reference_b'],
            qty_a, qty_b, diff_qty, r['match_confidence']
        ]
        fill = SEVERITY_STYLES.get(r['severity'], PatternFill())

        for col_idx, val in enumerate(row_data, 1):
            cell = ws_detail.cell(row=row_idx, column=col_idx, value=val)
            cell.font = NORMAL_FONT
            cell.fill = fill
            cell.alignment = CENTER
            cell.border = BORDER

    # Auto-width
    for col in range(1, len(headers) + 1):
        ws_detail.column_dimensions[ws_detail.cell(row=1, column=col).column_letter].width = 18

    # Save
    filename = f"BOM差异报告_{task_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    save_path = os.path.join(
        os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..')),
        'reports', filename
    )
    os.makedirs(os.path.dirname(save_path), exist_ok=True)
    wb.save(save_path)

    # Update task summary
    db.execute(
        'UPDATE comparison_task SET summary=? WHERE id=?',
        (f'差异总数：{total} 条', task_id)
    )

    return save_path


# ======================== 返工变更单导出 ========================

REWORK_REMOVE_FILL = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
REWORK_INSTALL_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
REWORK_ADJUST_FILL = PatternFill(start_color='FFE5CC', end_color='FFE5CC', fill_type='solid')

REWORK_REMOVE_FONT = Font(bold=True, color='B71C1C', size=10)
REWORK_INSTALL_FONT = Font(bold=True, color='1B5E20', size=10)
REWORK_ADJUST_FONT = Font(bold=True, color='E65100', size=10)


def generate_rework_order_excel(task_id):
    """Generate a rework change order Excel for production line execution.

    Converts BOM comparison results into actionable rework instructions:
      - 需拆除 (remove): materials in current unit but NOT in target
      - 需补装 (install): materials in target but NOT in current unit
      - 用量调整 (adjust): same PN, different quantity
    """
    task = db.query_one('SELECT * FROM comparison_task WHERE id=?', (task_id,))
    if not task:
        return None
    task = dict(task)

    source_bom = db.query_one('SELECT bom_name FROM bom_header WHERE id=?', (task['source_bom_id'],))
    target_bom = db.query_one('SELECT bom_name FROM bom_header WHERE id=?', (task['target_bom_id'],))

    results = db.query(
        'SELECT * FROM comparison_result WHERE task_id=? ORDER BY diff_type, id',
        (task_id,)
    )

    remove_items = []
    install_items = []
    adjust_items = []

    for r in results:
        r = dict(r)
        if r['diff_type'] == 'added':
            remove_items.append(r)
        elif r['diff_type'] == 'removed':
            install_items.append(r)
        elif r['diff_type'] == 'modified' and r['diff_category'] == 'quantity':
            adjust_items.append(r)

    wb = Workbook()

    # ================ Sheet 1: 返工概览 ================
    ws = wb.active
    ws.title = '返工概览'

    overview = [
        ('', ''),
        ('返工变更单', ''),
        ('', ''),
        ('当前实物 BOM', target_bom['bom_name'] if target_bom else ''),
        ('返工目标 BOM', source_bom['bom_name'] if source_bom else ''),
        ('比对任务', task.get('task_name', '')),
        ('生成时间', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
        ('', ''),
        ('--- 变更汇总 ---', ''),
        ('需拆除项数', len(remove_items)),
        ('需补装项数', len(install_items)),
        ('用量调整项数', len(adjust_items)),
        ('变更总数', len(remove_items) + len(install_items) + len(adjust_items)),
    ]

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 50

    for row_idx, (key, val) in enumerate(overview, 1):
        if key.startswith('---'):
            ws.cell(row=row_idx, column=1, value=key).font = Font(bold=True, size=11, color='475569')
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
        elif key == '返工变更单':
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
            cell = ws.cell(row=row_idx, column=1, value=key)
            cell.font = Font(bold=True, size=16, color='0F172A')
            cell.alignment = CENTER
        elif key == '':
            pass
        else:
            ws.cell(row=row_idx, column=1, value=key).font = Font(bold=True, size=10)
            ws.cell(row=row_idx, column=2, value=val).font = NORMAL_FONT

    # ================ Sheet 2: 需拆除清单 ================
    ws_remove = wb.create_sheet('需拆除清单')
    _write_rework_section(ws_remove, '需拆除物料清单',
                          ['序号', '物料编码', '物料名称', '需拆除用量', '单位', '位号'],
                          remove_items, REWORK_REMOVE_FILL, REWORK_REMOVE_FONT,
                          pn_field='part_number_b', name_field='part_name_b',
                          qty_field='quantity_b', ref_field='reference_b')

    # ================ Sheet 3: 需补装清单 ================
    ws_install = wb.create_sheet('需补装清单')
    _write_rework_section(ws_install, '需补装物料清单',
                          ['序号', '物料编码', '物料名称', '需补装用量', '单位', '位号'],
                          install_items, REWORK_INSTALL_FILL, REWORK_INSTALL_FONT,
                          pn_field='part_number_a', name_field='part_name_a',
                          qty_field='quantity_a', ref_field='reference_a')

    # ================ Sheet 4: 用量调整清单 ================
    ws_adjust = wb.create_sheet('用量调整清单')
    _write_rework_section(ws_adjust, '用量调整清单',
                          ['序号', '物料编码', '物料名称', '当前用量', '目标用量', '单位', '位号'],
                          adjust_items, REWORK_ADJUST_FILL, REWORK_ADJUST_FONT,
                          pn_field='part_number_a', name_field='part_name_a',
                          qty_field='quantity_a', ref_field='reference_a',
                          extra_field='new_value')

    filename = f"返工变更单_{task_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    save_path = os.path.join(
        os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..')),
        'reports', filename
    )
    os.makedirs(os.path.dirname(save_path), exist_ok=True)
    wb.save(save_path)
    return save_path


def _write_rework_section(ws, title, headers, items, header_fill, header_font,
                          pn_field='part_number_b', name_field='part_name_b',
                          qty_field='quantity_b', ref_field='reference_b',
                          extra_field=None):
    """Write a rework section sheet with consistent formatting."""
    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    cell = ws.cell(row=1, column=1, value=title)
    cell.font = Font(bold=True, size=14, color='0F172A')
    cell.alignment = CENTER

    row_idx = 3

    # Column headers
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=row_idx, column=col_idx, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER

    row_idx = 4

    if not items:
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(headers))
        c = ws.cell(row=row_idx, column=1, value='（无）')
        c.font = Font(size=10, color='94A3B8')
        c.alignment = CENTER
        return

    for seq, it in enumerate(items, 1):
        pn = it.get(pn_field, '')
        name = it.get(name_field, '')
        qty = _fmt_qty(it.get(qty_field, ''))
        ref = it.get(ref_field, '')

        if extra_field:
            # 用量调整：展示 当前用量 / 目标用量 两列
            row_data = [seq, pn, name, qty, _fmt_qty(it.get(extra_field, '')), '', ref]
        else:
            row_data = [seq, pn, name, qty, '', ref]

        for col_idx, val in enumerate(row_data, 1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.font = NORMAL_FONT
            c.fill = header_fill
            c.alignment = CENTER if col_idx in (1, 5, 6) else LEFT_WRAP
            c.border = BORDER
        row_idx += 1

    # Column widths
    widths = [8, 28, 38, 12, 12, 8, 22]
    for i, w in enumerate(widths[:len(headers)]):
        col_letter = ws.cell(row=3, column=i + 1).column_letter
        ws.column_dimensions[col_letter].width = w
